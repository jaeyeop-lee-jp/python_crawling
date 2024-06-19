#########################################################################################################
# Worknet 채용공고 수집 프로그램
# 사용 모듈 선언
#########################################################################################################

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import pandas  as pd    
import numpy  
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
from fake_useragent import UserAgent

import time, sys, math, os, random


#########################################################################################################
# Worknet 정보 수집 시작 - 기간별 수집 
#########################################################################################################

query_txt = "로봇"
f_dir = "c:/korea_robot/"

chrome_path = "c:/korea_robot/chromedriver.exe"

options = Options()
ua = UserAgent()
userAgent = ua.random
options.add_argument("user-agent={userAgent}")

# 크롬드라이버 설정

options = webdriver.ChromeOptions()
#options.add_argument('headless')
#options.add_argument('window-size=1920x1080')
options.add_argument("--disable-gpu")
args = ["hide_console", ]
#driver = webdriver.Chrome(path)
#driver = webdriver.Chrome(chrome_path,options=options,service_args=args)
driver = webdriver.Chrome(options=options, executable_path=r'c:/korea_robot/chromedriver.exe')


driver.get('http://www.work.go.kr')
driver.maximize_window()

time.sleep(2)


#########################################################################################################
# Worknet 정보 수집 시작 - 검색 기간 설정
#########################################################################################################

# 검색창에 입력 받은 검색어를 넣고 검색을 실행합니다
element = driver.find_element_by_id("topQuery")
element.send_keys(query_txt)
element.send_keys("\n")

driver.find_element_by_xpath('//*[@id="contents"]/div/div[1]/div[1]/div[2]/div[3]/a').click()
search_range = int(input('조회 기간 선택 - 전체(1), 오늘(2), 최근3일(3), 최근 1개뤌(4) : '))

# 상세검색 클릭
driver.find_element_by_xpath('//*[@id="contents"]/div/div[1]/div[1]/div[2]/div[1]/div/a').click()
time.sleep(2)
driver.execute_script("window.scrollTo(0, 700)") 
time.sleep(2)
driver.find_element_by_xpath('//*[@id="srcFrm"]/div[1]/div[3]/button[1]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="b_siteClcdWORK"]').click()
time.sleep(2)


if search_range == 1 :
    driver.find_element_by_xpath('//*[@id="termSearchGbn0"]').click()
elif search_range == 2 :
    driver.find_element_by_xpath('//*[@id="termSearchGbn1"]').click()
elif search_range == 3 :
    driver.find_element_by_xpath('//*[@id="termSearchGbn2"]').click()
elif search_range == 4 :
    driver.find_element_by_xpath('//*[@id="termSearchGbn5"]').click()
else :
    print("유효한 번호를 입력하세요")


# 검색 버튼 클릭
driver.find_element_by_xpath('//*[@id="srcFrm"]/div[3]/div[4]/button').click()

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

search_cnt = soup.find('em', class_='futura').get_text()
cnt = int(search_cnt)
page_cnt = math.ceil(cnt / 10)

print('%s 검색어로 검색된 결과는 총 %s 건입니다' %(query_txt, search_cnt))


#########################################################################################################
# 전체 구인광고의 URL 주소를 먼저 추출하여 목록을 만든 후 한건씩 상세 정보를 추출
#########################################################################################################

all_url = [ ]
url_no = 1

for x in range(1,page_cnt + 1) :           

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')       
    list_1 = soup.find_all('div', class_='cp-info-in')

    for a in list_1 :
        url_1 = 'https://www.work.go.kr'
        url_2 = a.find('a')['href']
        full_url = url_1 + url_2
        #print(url_no,':',full_url)
        all_url.append(full_url)

        # print(full_url)
        
        url_no += 1
        
        if url_no > cnt :
            break
                    
    if x == page_cnt :
        print('요청하신 URL의 수집이 완료되었습니다')

    x += 1

    if x > page_cnt :
        break

    try :
        driver.find_element_by_link_text("""%s""" %x).click() # 다음 페이지번호 클릭
    except :
        continue


driver.close()



#########################################################################################################
# 추출한 URL 주소를 이용 새 브라우저 오픈하여 자료 수집
#########################################################################################################

# 구인광고 목록 만들기
no2=  [ ]            #1.번호
company_name2=[ ]    #2.기업명
upjong2 = [ ]        #3.업종
work2 = [ ]          #4.직무내용
carrier2 = [ ]       #5.경력여부
school2 = [ ]        #6.학력여부
normal2 = [ ]        #7.고용형태
many2 = [ ]          #8.모집인원
pay2 = [ ]           #9.급여
area2 = [ ]          #10.근무지역
time2 = [ ]          #11.근무시간
url2 = [ ]           #12.원본 URL

no = 1           # 게시글 번호 초기값


# 크롬 드라이버를 사용해서 웹 브라우저를 실행합니다.
options = webdriver.ChromeOptions()
#options.add_argument('headless')
#options.add_argument('window-size=1920x1080')
options.add_argument("--disable-gpu")
args = ["hide_console", ]
#driver = webdriver.Chrome(path)
driver2 = webdriver.Chrome(chrome_path, options=options,  service_args=args)



for x in range(0, len(all_url)) : 
    #print(all_url[x])
    #print(userAgent)
    driver2.get(all_url[x])
    time.sleep(1)

    html = driver2.page_source
    soup = BeautifulSoup(html, 'html.parser')       

    # 1.게시글 번호
    no2.append(str(no))
    
    # 2.회사명
    company= soup.find('div', class_='right').find('div', class_='info').find('ul').find_all('li')
    try :
        company_name = company[0].find('div').get_text().strip()
    except :
        company_name = "회사정보가 없습니다."
        company_name2.append(company_name)
    else :
        company_name2.append(company_name)

    # 3.업종
    try :
        upjong = company[1].find('div').get_text().strip()
    except :
        upjong = "업종정보가 없습니다."    
        upjong2.append(upjong)
    else :
        upjong2.append(upjong)


    # 4. 직무내용
    try : 
        work_lst1 = soup.find('div', class_='left')
        work = work_lst1.find('p', class_='tit').get_text().strip()
    except ValueError:
        work = "직무정보가 없습니다."
        work2.append(work)
    else :
        work2.append(work)

    # 5. 경력 
    work_lst2 = soup.find('div', class_='left').find_all('div', class_='info')
    carrier_imsi1 = work_lst2[0].find('div', class_='cont').find('ul').find_all('li')
    try :
        carrier_imsi2 = carrier_imsi1[0].find('span').get_text().strip()
        carrier = ' '.join(carrier_imsi2.split())
    except : 
        carrier = "경력정보가 없습니다."
        carrier2.append(carrier)
    else :
        carrier2.append(carrier)

    # 6. 학력
    try :
        school = carrier_imsi1[1].find('span').get_text().strip()
    except :
        school = "학력정보가 없습니다."        
        school2.append(school)
    else :
        school2.append(school)
    
    # 07. 고용형태
   
    try :
        normal_imsi1 = work_lst2[1].find('div', class_='cont').find('ul').find('li')
        normal = normal_imsi1.find('span').get_text().strip()
    except :
        normal = "고용형태 정보가 없습니다."
        normal2.append(normal)
    else :
        normal2.append(normal)


    # 10. 근무지역
    job_work_imsi1 = work_lst2[0].find_all('div', class_='column')
    job_work_imsi2 = job_work_imsi1[1].find('div', class_='cont').find('ul').find_all('li')
        
    try :
        job_work_imsi3 = job_work_imsi2[0].find('span').get_text().strip()
    except :
        job_work_imsi3 = "근무지 정보가 없습니다."
        area2.append(job_work_imsi3)
    else :
        area2.append(job_work_imsi3)

    # 09. 임금
    try :
        job_work_imsi4 = job_work_imsi2[1].find('span').get_text().strip()
        job_work_imsi5 = ' '.join(job_work_imsi4.split())
    except :
        job_work_imsi5 = "임금 정보가 없습니다."
        pay2.append(job_work_imsi4)
    else :
        pay2.append(job_work_imsi4)

    # 04. 직무
    try :
        careers_imsi1 = soup.find_all('div', class_='careers-table v1 center mt20')
        careers_imsi2 = careers_imsi1[1].find('table').find('tbody').find('tr').find('td')
        careers_imsi3 = careers_imsi2.get_text().strip()
        work2.append(careers_imsi3)
    except :
        careers_imsi3 = "직무 정보가 없습니다."
        work2.append(careers_imsi3)        

    # 08. 모집인원
    try :
        inwon_imsi1 = careers_imsi1[0].find('table').find('tbody').find('tr').find_all('td')
        inwon_imsi2 = inwon_imsi1[3].get_text().replace('입사지원 현황통계', '').strip()
        many = ' '.join(inwon_imsi2.split())
        many2.append(many)
    except :
        many = "모집인원 정보가 없습니다."
        many2.append(many)



    #########################################################################################################
    # GUI 출력을 위한 부분 
    #########################################################################################################

    print("2. 회사명 : {} " .format(company_name))
    print("3. 업종 : {} " .format(upjong))
    print("4. 직무 : {} " .format(careers_imsi3))
    print("5. 경력 : {} " .format(carrier))
    print("6. 학력 : {} " .format(school))
    print("7. 고용형태 : {} " .format(normal))
    print("8. 모집인원 : {} " .format(inwon_imsi2))
    print("9. 임금 : {} " .format(job_work_imsi5))
    print("10. 근무지역 : {} " .format(job_work_imsi3))
    # print("11. 근무시간 : " )
    print("12. 원본URL : {}" .format(all_url[x]))
    print('\n')


    time.sleep(random.randrange(10,50))

    if no > cnt :
        break
   
    no += 1

driver2.close()        

#########################################################################################################
# 출력 결과물 저장하기
#########################################################################################################


df = pd.DataFrame()
# df['번호']=no2
df['회사명'] = pd.Series(company_name2)
df['업종'] = pd.Series(upjong2)
df['직무내용'] = pd.Series(work2)
df['경력여부'] = pd.Series(carrier2) 
df['학력사항'] = pd.Series(school2)
df['고용형태'] = pd.Series(normal2)
df['모집인원'] = pd.Series(many2)
df['급여'] = pd.Series(pay2)
df['근무지역'] = pd.Series(area2)
# df['근무시간'] = time2 
# df['원본URL'] = url2 


                 
xlfile_path = r'c:/korea_robot/data/korea_robot_worknet.xlsx'
imsi_df = pd.read_excel(xlfile_path)

wb = load_workbook(xlfile_path)
writer = pd.ExcelWriter(xlfile_path, engine='openpyxl')

sheet = wb['worknet']

imsi_df.to_excel(writer, 'worknet', index = False)
df.to_excel(writer, 'worknet', startrow = sheet.max_row, index = False, header = None)

writer.save()
writer.close()
