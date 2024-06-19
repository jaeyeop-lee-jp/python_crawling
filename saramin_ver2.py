#########################################################################################################
# Saramin 채용공고 수집 프로그램
# 사용 모듈 선언
#########################################################################################################

from bs4 import BeautifulSoup 
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
import pandas  as pd    
import numpy  
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

from datetime import datetime, timedelta
import requests
import os.path

import time, os, math, re, sys, random


def xpath_click(xpath_name):
    driver.find_element_by_xpath(xpath_name).click()
    time.sleep(1)

def id_click(id_name):
    driver.find_element_by_id(id_name).click()
    time.sleep(1)

def class_click(cl_name):
    driver.find_element_by_class_name(cl_name).click()
    time.sleep(1)


#########################################################################################################
# Saramin 정보 수집 시작 - 기간별 수집 
#########################################################################################################

# 주요 사용자 입력 번수 선언
query_txt = "로봇"  
f_dir = "c:/korea_robot/"  

chrome_path = "c:/korea_robot/chromedriver.exe"


# 웹사이트 접속(크롬드라이버)
url = 'http://www.saramin.co.kr/zf_user/search/recruit?search_area=main&search_done=\
    y&search_optional_item=n&searchType=default_mysearch&searchword='+ query_txt + '\
        &recruitPage=1&recruitSort=relation&recruitPageCount=40&inner_com_type=&comp\
            any_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply='

options = Options()
ua = UserAgent()
userAgent = ua.random
options.add_argument("user-agent={userAgent}")

options = webdriver.ChromeOptions()
#options.add_argument('headless')
#options.add_argument('window-size=1920x1080')
options.add_argument("--disable-gpu")
args = ["hide_console", ]
#driver = webdriver.Chrome(path)
#driver = webdriver.Chrome(chrome_path,options=options,service_args=args)
driver = webdriver.Chrome(options=options, executable_path=r'c:/korea_robot/chromedriver.exe')

driver.get(url)
time.sleep(2)


###################################################################################
#  공고 100개 조회 조회
###################################################################################

# now_time = datetime.now()

driver.refresh()

# html = driver.page_source
# soup = BeautifulSoup(html, 'html.parser')

# search_cnt_1 = soup.find('span' , class_= "cnt_result").get_text().replace(",","")
# search_cnt_2 = int(re.findall("\d+", search_cnt_1)[0])
# page_cnt = math.ceil(search_cnt_2 / 100)  # 크롤링 할 페이지 수 
# # print('조회된 공고는 {} 건 입니다' .format(search_cnt_2))

xpath_click('//*[@id="recruit_info"]/div[2]/div/div[2]/button')
xpath_click('//*[@id="recruit_info"]/div[2]/div/div[2]/div/ul/li[3]/button')
xpath_click('//*[@id="recruit_info"]/div[2]/div/div[3]/button')
xpath_click('//*[@id="recruit_info"]/div[2]/div/div[3]/div/ul/li[1]/button')
time.sleep(2)


# ###################################################################################
# #  신규 공고 검출 여부 확인
# ###################################################################################  

lst_tit_index = []

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
content_list = soup.find_all('div', class_='item_recruit')

for i in content_list :
    tit_idx = i['value']
    lst_tit_index.append(tit_idx)

print('공고목록 조회결과 : {} ' .format(lst_tit_index))
        

joblist_in_file = pd.read_excel("c:/korea_robot/data/korea_robot_saramin.xlsx",\
    header=0, usecols='A', sheet_name='saramin')


# 데이터 프레임을 리스트로 전환

# lst_tit_index_str = list(map(str, lst_tit_index))

lst_tit_index_data1 = joblist_in_file.values.tolist()
lst_tit_index_data2 = (sum(lst_tit_index_data1, []))
lst_tit_index_data3 = list(map(str, lst_tit_index_data2))

# 저장된 데이터와 수집된 데이터를 비교
job_tit_list_result = list(set(lst_tit_index) - set(lst_tit_index_data3))

if  len(job_tit_list_result) != 0 :
    print('신규 정보가 총 {} 건 있습니다.' .format(len(job_tit_list_result)))
    time.sleep(5)
else :
    print('신규 정보가 없습니다.')


time.sleep(10)


# 저장용 리스트 선언
# lst_no=[]             # 게시글 번호 컬럼
lst_code =[]            # 게시물 고유코드
lst_cname=[]            # 회사명 컬럼
# lst_upjong=[]           # 업종
lst_jikjong=[]          # 모집직종
lst_jikmu=[]            # 직무
lst_spec=[]             # 경력
lst_univ=[]             # 학력
lst_ypay=[]             # 임금
# lst_goyong=[]           # 고용형태
# lst_mojipinwon=[]       # 모집인원
lst_area=[]             # 근무지역
lst_jeoupsu_s=[]         # 접수시작
lst_jeoupsu_e=[]         # 접수종료
lst_link=[]             # 원본 url

content_list =[]
no = 1           # 게시글 번호 초기값


########################################################################################################
# 게시물 정보 추출, 신규정보가 있을 경우 채용공고 수집
# 신규 정보 존재 여부 비교 후 반복문 실행
#########################################################################################################

if len(job_tit_list_result) != 0 :
    
    print('자료 수집을 시작합니다')

    driver.refresh()
    time.sleep(3)

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    # search_cnt_1 = soup.find('span' , class_= "cnt_result").get_text().replace(",","")
    # search_cnt_2 = int(re.findall("\d+", search_cnt_1)[0])
    # page_cnt = math.ceil(search_cnt_2 / 80)  # 크롤링 할 페이지 수
    # print(page_cnt)

    # 00건 보기 클릭
    xpath_click('//*[@id="recruit_info"]/div[2]/div/div[2]/button')
    xpath_click('//*[@id="recruit_info"]/div[2]/div/div[2]/div/ul/li[3]/button')
    xpath_click('//*[@id="recruit_info"]/div[2]/div/div[3]/button')
    xpath_click('//*[@id="recruit_info"]/div[2]/div/div[3]/div/ul/li[1]/button')


    for j_idx in job_tit_list_result :
        
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
                        
        print('{} 번 공고 수집' .format(j_idx))
        content = soup.find('div', value = j_idx )
        lst_code.append(j_idx)

        # 기본정보 추출 : 회사명(caname)
        try : 
            cname = content.find('strong', class_= 'corp_name').find('span').get_text()
        except :
            cname = "정보가 없습니다"
            lst_cname.append(cname)
        else :
            lst_cname.append(cname)

        # 기본정보 추출 : 직종(jikjong), 직무(jikmu)
        try :
            jikjongs = content.find('div', class_= 'job_sector').find_all('a')
            jikjong1 = [i.get_text().strip() for i in jikjongs]
            jikjong = ','.join(jikjong1)
        except :
            jikjong = "정보가 없습니다"
            lst_cname.append(jikjong)
        else :
            lst_cname.append(jikjong)
        
        try :
            jikmu = content.find('h2', class_= 'job_tit').find('a').get_text()
        except :
            jikmu = "정보가 없습니다"
            lst_cname.append(jikmu)
        else :
            lst_cname.append(jikmu)


        # 기본정보 추출 : 근무지역(area), 경력(spec), 학력(univ), 고용형태(goyong)
        job_info = content.find('div', class_= 'job_condition').find_all('span')
                
        try :
            areas = job_info[0].find_all('a')
            area1 = [i.get_text().strip() for i in areas]
            area = ','.join(area1)
        except :
            area = "정보가 없습니다"
            lst_cname.append(area)
        else :
            lst_cname.append(area)

        try :
            spec = job_info[1].get_text().replace('↑', ' 이상')
        except :
            spec = "정보가 없습니다."
            lst_cname.append(spec)
        else :
            lst_cname.append(spec)

        try :
            univ = job_info[2].get_text().replace('↑', ' 이상')
        except :
            univ = "정보가 없습니다"
            lst_cname.append(univ)
        else :
            lst_cname.append(univ)
                
        
        # goyong = job_info[3].get_text()


        # 기본정보 추출 : 게시물 링크
        try : 
            link = 'http://www.saramin.co.kr/zf_user/jobs/relay/view?view_type=search&rec_idx='+j_idx+'&location=ts&searchword='+query_txt+'&searchType=default_mysearch&paid_fl=n'
        except :
            link = "정보가 없습니다."
            lst_cname.append(link)
        else :
            lst_cname.append(link)


        #######################################################################################
        # 상세정보 추출을 위한 개별 게시물 클릭, 새 탭으로 오픈, 브라우저 핸들 1번
        #######################################################################################

        # 크롬 드라이버를 사용해서 웹 브라우저를 실행합니다.
        options = webdriver.ChromeOptions()
        #options.add_argument('headless')
        #options.add_argument('window-size=1920x1080')
        options.add_argument("--disable-gpu")
        args = ["hide_console", ]
        #driver = webdriver.Chrome(path)
        driver2 = webdriver.Chrome(chrome_path, options=options,  service_args=args)

        # driver.execute_script('window.open("about:blank", "_blank");')
        time.sleep(5)
        
        driver2.get(link)
        time.sleep(10)
        # driver.switch_to_window(driver.window_handles[1])

        html = driver2.page_source
        soup = BeautifulSoup(html, 'html.parser')

        # 데이터 추출 : 업종정보
        # try :
        #     upjong_imsi = soup.find('dt', text='업종')
        #     upjong_is =upjong_imsi.find('dd').get_text()
        # except :
        #     upjong = "업종 정보가 없습니다."
        #     lst_upjong.append(upjong)
        # else :
        #     upjong = str(upjong_is)
        #     lst_upjong.append(upjong)


        # 데이터 추출 : 임금정보
        imsi = soup.find('div', class_= 'cont').find_all('div', class_=  'col')
                
        try :
            ypay1 = imsi[1].find('dl' > 'dd').get_text().replace('급여', '').replace('상세보기', '').replace('닫기', '').strip()
            ypay = ' '.join(ypay1.split())
        except :
            ypay = "임금 정보가 없습니다."
            lst_cname.append(ypay)
        else :
            lst_cname.append(ypay)


        # 데이터 추출 : 접수기간
        try :
            jeoupsu_imsi = soup.find('dl', class_= 'info_period').find('dd')
            jeoupsu_s = jeoupsu_imsi.get_text()
        except :
            jeoupsu_s = "시작날짜가 없습니다"
            lst_cname.append(jeoupsu_s)
        else :
            lst_cname.append(jeoupsu_s)

        try :
            jeoupsu_imsi = soup.find('dl', class_= 'info_period').find_all('dd')
            jeoupsu_e = jeoupsu_imsi[1].get_text()
            if jeoupsu_e != "채용시" :
                lst_cname.append(jeoupsu_e)
        except :
            jeoupsu_e = "정보가 없습니다"
            lst_cname.append(jeoupsu_e)
        else :
            lst_cname.append(jeoupsu_e)
        

        # 프린트 ,  차후 GUI 만들기 위한 화면 출력용

        print('고유코드 : {}' .format(j_idx))
        print('회사명 : {}' .format(cname))
        # print('업종 : {} ' .format(upjong))
        print('직종 : {}' .format(jikjong))
        print('직무 : {}' .format(jikmu))
        print('경력 : {}' .format(spec))
        print('학력 : {}' .format(univ))
        # print(goyong)
        # print(mojipinwon)     # 모집인원 수집 불가, 해당정보 없음
        print('임금 : {}' .format(ypay))
        print('근무지역 : {}' .format(area))
        print('접수시작 : {}' .format(jeoupsu_s))
        print('접수종료 : {}' .format(jeoupsu_e))
        print('원본 URL : {}' .format(link))
        print("\n")


        driver2.close()
        time.sleep(5)
        
        driver.switch_to_window(driver.window_handles[0])

   


#########################################################################################################
# 출력 결과물 저장하기
#########################################################################################################

df = pd.DataFrame()
# df.astype('object').dtypes

# saramin['번호']=no2
df['고유코드']=pd.Series(lst_code)
df['회사명']=pd.Series(lst_cname)
# saramin['업종']=pd.Series(lst_upjong)
df['모집직종']=pd.Series(lst_jikjong)
df['직무']=pd.Series(lst_jikmu)
df['경력']=pd.Series(lst_spec)
df['학력']=pd.Series(lst_univ)
# saramin['고용형태']=pd.Series(lst_goyong)
# saramin['모집인원']=pd.Series(lst_mojipinwon)
df['임금']=pd.Series(lst_ypay)
df['근무지역']=pd.Series(lst_area)
df['접수시작']=pd.Series(lst_jeoupsu_s)
df['접수종료']=pd.Series(lst_jeoupsu_e)
df['원본URL']=pd.Series(lst_link)



xlfile_path = r'c:/korea_robot/data/korea_robot_saramin.xlsx'
imsi_df = pd.read_excel(xlfile_path)
# imsi_df.astype('object').dtypes

wb = load_workbook(xlfile_path)
writer = pd.ExcelWriter(xlfile_path, engine='openpyxl')

sheet = wb['saramin']

imsi_df.to_excel(writer, 'saramin', index = False)
df.to_excel(writer, 'saramin', startrow = sheet.max_row, index = False, header = None)

writer.save()
writer.close()
driver.close()